"""
XLSX exporter for Excel spreadsheets.
"""
from typing import Dict
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from app.exporters.base import BaseExporter

logger = logging.getLogger(__name__)


class XLSXExporter(BaseExporter):
    """Export OCR results to XLSX format."""

    def export(self, data: Dict, output_path: str) -> str:
        """
        Export to XLSX file.

        Args:
            data: OCR result data
            output_path: Output file path

        Returns:
            Path to created file
        """
        self._ensure_output_dir(output_path)

        try:
            wb = Workbook()

            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            # Add text sheet if there's text content
            text_data = data.get("text", {})
            full_text = text_data.get("full_text", "")

            if full_text:
                ws_text = wb.create_sheet("Текст")
                ws_text['A1'] = full_text
                ws_text['A1'].alignment = Alignment(wrap_text=True, vertical='top')
                ws_text.column_dimensions['A'].width = 100

            # Add table sheets
            tables = data.get("tables", [])
            for idx, table_data in enumerate(tables, 1):
                cells = table_data.get("cells", [])
                if not cells:
                    continue

                ws = wb.create_sheet(f"Таблица {idx}")

                # Write data
                for row_idx, row_data in enumerate(cells, 1):
                    for col_idx, cell_value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=str(cell_value))

                        # Style header row
                        if row_idx == 1:
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal='center')

                # Apply merged cells
                merged_regions = table_data.get("merged_regions", [])
                for merge in merged_regions:
                    start_row = merge['row'] + 1  # openpyxl is 1-indexed
                    start_col = merge['col'] + 1
                    end_row = start_row + merge['rowspan'] - 1
                    end_col = start_col + merge['colspan'] - 1

                    # Only merge if it's actually a merged region
                    if end_row > start_row or end_col > start_col:
                        try:
                            ws.merge_cells(
                                start_row=start_row, start_column=start_col,
                                end_row=end_row, end_column=end_col
                            )
                        except Exception as e:
                            logger.warning(f"Could not merge cells ({start_row},{start_col})-({end_row},{end_col}): {e}")

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # If no sheets were created, add a default one
            if not wb.sheetnames:
                wb.create_sheet("Данные")
                wb["Данные"]['A1'] = "Нет данных для экспорта"

            wb.save(output_path)
            logger.info(f"Exported to XLSX: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error exporting to XLSX: {e}")
            raise
