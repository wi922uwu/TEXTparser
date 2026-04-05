"""Export tables to XLSX format."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from backend.models.schemas import TableRegion


def export_xlsx(
    table_regions: list[TableRegion],
    output_path: Path,
) -> Path:
    """Export extracted tables as XLSX with formatting."""
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(
        start_color="D9E2F3",
        end_color="D9E2F3",
        fill_type="solid",
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for table_idx, table_data in enumerate(table_regions):
        if not table_data.cells:
            continue

        ws = wb.create_sheet(title=f"Table_{table_idx + 1}")

        for row_idx, row_cells in enumerate(table_data.cells):
            is_header = row_idx == 0
            for col_idx, cell_value in enumerate(row_cells, 1):
                cell = ws.cell(
                    row=row_idx + 1,
                    column=col_idx,
                    value=str(cell_value),
                )
                cell.alignment = Alignment(wrap_text=True)
                cell.border = thin_border

                if is_header:
                    cell.font = header_font
                    cell.fill = header_fill

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted

    wb.save(str(output_path))
    return output_path
